from flask import Flask, render_template, request, flash, redirect, url_for
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_
import pandas as pd
from datetime import datetime

app = Flask(__name__)

app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///database_file.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.secret_key = "sdf"
db = SQLAlchemy(app)


class User(db.Model):
    user_id = db.Column(db.Integer, primary_key=True)
    ID = db.Column(db.Integer, unique=True, nullable=False)
    NAME = db.Column(db.String(50), nullable=False)
    SURNAME = db.Column(db.String(50), nullable=False)
    NUMBER = db.Column(db.Integer,  nullable=False)
    PHONE = db.Column(db.Integer, unique=True, nullable=False)
    TIME = db.Column(db.DateTime, default=datetime.now)  # Add a new column for the creation time


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        NAME = request.form.get("name")
        ID = request.form.get("ID")
        SURNAME = request.form.get("surname")
        NUMBER = request.form.get("number")
        PHONE = request.form.get("tel")

        if not NAME or not ID or not SURNAME or not NUMBER or not PHONE:
            flash("Please fill out all the required fields.⚠️")
        else:
            # Check if the user already exists in the database
            id_exists = User.query.filter_by(ID=ID).first()
            num_exists = User.query.filter_by(PHONE=PHONE).first()

            if id_exists:
                flash("ID already exists. Please choose a different ID.⚠️")
                return redirect(url_for("index"))

            if num_exists:
                flash("Phone number already exists. Please choose a different phone number.⚠️")
                return redirect(url_for("index"))

            # Validate ID length
            if len(str(ID)) != 11:
                flash("ID must be 11 digits long.⚠️")
                return redirect(url_for("index"))

            # Validate phone number starting from 5 and be 9 characters long
            if not str(PHONE).startswith("5") or len(str(PHONE)) != 9:
                flash("Phone number must start with 5 and be 9 characters long.⚠️")
                return redirect(url_for("index"))

            new_user = User(ID=ID, NAME=NAME, SURNAME=SURNAME, NUMBER=NUMBER, PHONE=PHONE)
            db.session.add(new_user)
            db.session.commit()

            flash("Information saved.✅")
            save_to_excel()
            return redirect(url_for("index"))

    return redirect(url_for("index"))


@app.route("/users")
def display():
    search_query = request.args.get("search")
    if search_query:
        users = User.query.filter(
            or_(
                User.ID.like(f"%{search_query}%"),
                User.PHONE.like(f"%{search_query}%"),
            )
        ).all()
    else:
        users = User.query.all()

    return render_template("users.html", users=users, search_query=search_query if search_query else "")


@app.route("/delete/<id>", methods=["POST"])
def delete_user(id):
    user = User.query.filter_by(user_id=id).first()
    if user:
        db.session.delete(user)
        db.session.commit()
        flash("User deleted successfully.✅")
        save_to_excel()
    else:
        flash("User not found.⚠️")
    return redirect(url_for("display"))


import openpyxl
from openpyxl.utils import get_column_letter

def save_to_excel():
    users = User.query.all()
    data = {
        "ID": [user.ID for user in users],
        "Name": [user.NAME for user in users],
        "Surname": [user.SURNAME for user in users],
        "Number": [user.NUMBER for user in users],
        "Phone": [user.PHONE for user in users],
        "Day": [user.TIME.strftime('%d/%m') for user in users],
        "Time": [user.TIME.strftime('%H:%M') for user in users]
    }

    df = pd.DataFrame(data)
    df.to_excel("user_data.xlsx", index=False)

    # Adjust column width to 100%
    workbook = openpyxl.load_workbook("user_data.xlsx")
    worksheet = workbook.active

    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save("user_data.xlsx")



# def run():
#     with app.app_context():
#         db.create_all()
#     app.run(host="0.0.0.0", debug=True)

# if __name__ == "__main__":
#     t = Thread(target=run)
#     t.start()
# if __name__ == "__main__":
with app.app_context():
    db.create_all()
app.run(host="0.0.0.0", port=5000 ,debug=True)